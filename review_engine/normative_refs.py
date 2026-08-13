#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
规范性引用文件自动审核工具 (a4guifan)
=====================================
输入 : 一份标准/规范类 .docx 文档
输出 : 1) 带 Word 批注(修订-批注)的 .docx
       2) 汇总所有问题与批注的 .xlsx

审核依据 : a4guifan_rules.json 中定义的 A4R001 ~ A4R015 共 15 条规则。

用法 :
    python audit_normative_refs.py input.docx
    python audit_normative_refs.py input.docx -o out.docx -x out.xlsx
    python audit_normative_refs.py input.docx --rules path/to/rules.json
"""

import argparse
import json
import os
import re
import sys
import datetime

from docx import Document
from docx.oxml.ns import qn
from docx.oxml import OxmlElement, parse_xml
from docx.text.paragraph import Paragraph
from docx.table import Table

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. 规则加载
# ---------------------------------------------------------------------------
def load_rules(rules_path):
    with open(rules_path, "r", encoding="utf-8") as f:
        rules = json.load(f)
    by_id = {r["rule_id"]: r for r in rules}
    return rules, by_id


# ---------------------------------------------------------------------------
# 2. 标准编号识别与分类
# ---------------------------------------------------------------------------
# 用于"规范性引用文件"章节内的宽松抽取(尽量抓到所有条目)
STD_LOOSE_RE = re.compile(
    r"(?:GB/Z|GB/T|GB|ISO/IEC|ISO|IEC|ITU-T|ITU-R|IEEE|ETSI|3GPP|OASIS|W3C|RFC|"
    r"DB\s*\d{2}(?:/T)?|T/[A-Z]+(?:\.[A-Z]+)*|"
    r"[A-Z]{2,5}(?:/[A-Z]{1,3})?)"
    r"\s*\d+(?:\.\d+)?(?:[-—]\d{4})?"
)

# 用于正文的严格抽取(避免把 NOTE/FIG 等误判为标准)
BODY_PREFIX_RE = re.compile(
    r"(?:GB/Z|GB/T|GB|ISO/IEC|ISO|IEC|ITU-T|ITU-R|IEEE|ETSI|3GPP|RFC|"
    r"DB\s*\d{2}(?:/T)?|T/[A-Z]+(?:\.[A-Z]+)*|"
    r"(?:YD|YS|YC|HG|HJ|JB|SJ|DL|NY|CJ|QC|QB|QJ|RB|SC|SH|SL|TB|WB|WH|WM|WW|XB|YB|YY|ZB|"
    r"GJB|HB|JC|JG|JT|LB|LD|LY|MH|NB|AQ|BB|CB|FZ|GA|GM|MT|SB|SN|SY|CJJ|JGJ)\s*/?[A-Z]?)"
    r"\s*\d+(?:\.\d+)?(?:[-—]\d{4})?"
)

# 把抓到的原始 token 解析成结构化信息(前缀只含字母/斜杠, 数字归到编号)
PARSE_RE = re.compile(
    r"^((?:3GPP)|(?:DB\s*\d{2}(?:/T)?)|[A-Za-z]+(?:/[A-Za-z]+)*)"
    r"\s*(\d+(?:\.\d+)?)(?:[-—](\d{4}))?$"
)

INTL_PREFIXES = {"ISO", "ISO/IEC", "IEC", "ITU-T", "ITU-R", "IEEE", "ETSI", "3GPP"}
INDUSTRY_CODES = {
    "YD", "YS", "YC", "HG", "HJ", "JB", "SJ", "DL", "NY", "CJ", "QC", "QB", "QJ",
    "RB", "SC", "SH", "SL", "TB", "WB", "WH", "WM", "WW", "XB", "YB", "YY", "ZB",
    "GJB", "HB", "JC", "JG", "JT", "LB", "LD", "LY", "MH", "NB", "AQ", "BB", "CB",
    "FZ", "GA", "GM", "MT", "SB", "SN", "SY", "CJJ", "JGJ",
}

# 排序层级: 国标0 -> 行业1 -> 地方2 -> 团体3 -> 国际4 -> 其他机构5 -> 其他文献6
TIER_ORDER = ["国家标准", "行业标准", "地方标准", "团体标准", "国际标准", "其他机构/文献", "其他文献"]


def parse_code(raw):
    """把抓到的原始编号解析为结构化字典, 无法解析返回 None。"""
    if raw is None:
        return None
    s = re.sub(r"\s+", "", raw)
    if "XXXX" in s.upper():
        return None  # 占位符交给 A4R015 处理
    m = PARSE_RE.match(s)
    if not m:
        return None
    prefix = m.group(1).upper()
    number = m.group(2)
    year = m.group(3)
    return {
        "raw": raw,
        "prefix": prefix,
        "number": number,
        "year": year,
        "is_intl": prefix in INTL_PREFIXES,
    }


def classify(prefix):
    """返回 (层级序号, 层级名称)。"""
    p = (prefix or "").upper()
    if p.startswith("GB"):
        return (0, "国家标准")
    if p.startswith("DB"):
        return (2, "地方标准")
    if p.startswith("T/"):
        return (3, "团体标准")
    if p in INTL_PREFIXES:
        return (4, "国际标准")
    if p in INDUSTRY_CODES or re.match(r"^[A-Z]{2,5}(/[A-Z])?$", p):
        return (1, "行业标准")
    return (5, "其他机构/文献")


def norm_base(code_text):
    """去掉空格/连字符/年代号, 用于跨章节-正文的一致性比对。"""
    s = re.sub(r"\s+", "", code_text).upper()
    s = re.sub(r"[-—]\d{4}.*$", "", s)
    return s


def extract_codes(text, strict=False):
    """抽取文本中所有标准编号, 返回 (原始串, 解析结果) 列表。"""
    rx = BODY_PREFIX_RE if strict else STD_LOOSE_RE
    out = []
    for m in rx.finditer(text):
        raw = m.group(0)
        out.append((raw, parse_code(raw)))
    return out


# ---------------------------------------------------------------------------
# 3. 文档结构解析
# ---------------------------------------------------------------------------
def iter_blocks(document):
    """按文档真实顺序返回块列表: (kind, obj), kind in {'p','tbl'}。"""
    blocks = []
    body = document.element.body
    for child in body:
        if child.tag == qn("w:p"):
            blocks.append(("p", Paragraph(child, document)))
        elif child.tag == qn("w:tbl"):
            blocks.append(("tbl", Table(child, document)))
    return blocks


def para_level(para):
    """返回标题层级(数字), 非标题返回 None。"""
    style = para.style
    name = (style.name or "") if style is not None else ""
    m = re.search(r"(\d+)", name)
    if "Heading" in name or "标题" in name:
        return int(m.group(1)) if m else 1
    pPr = para._p.pPr
    if pPr is not None:
        ol = pPr.find(qn("w:outlineLvl"))
        if ol is not None:
            val = ol.get(qn("w:val"))
            return int(val) + 1 if val is not None else 1
    return None


def is_heading(para):
    return para_level(para) is not None


def clean_heading_text(txt):
    """去掉标题前面的章节号, 如 '2 规范性引用文件' -> '规范性引用文件'。"""
    return re.sub(r"^\d+(\.\d+)*[.、]?\s*", "", txt).strip()


GUIDE_MARKERS = ["下列文件", "构成本文件", "规范性引用"]
NO_REF_TEXT = "本文件没有规范性引用文件"


def is_guide(text):
    t = text.strip()
    return ("下列文件" in t) and ("引用" in t or "条款" in t)


def is_no_ref(text):
    return NO_REF_TEXT in text


def find_reference_chapter(blocks):
    """定位'规范性引用文件'章节, 返回 (标题块索引, 章节内容块切片)。"""
    heading_idx = None
    for i, (kind, blk) in enumerate(blocks):
        if kind == "p" and is_heading(blk):
            t = clean_heading_text(blk.text)
            if "规范性引用文件" in t or "规范性引用文件" in blk.text.strip():
                heading_idx = i
                break
    if heading_idx is None:
        return None, []
    # 章节内容: 标题之后, 直到下一个标题
    content = []
    for j in range(heading_idx + 1, len(blocks)):
        kind, blk = blocks[j]
        if kind == "p" and is_heading(blk):
            break
        content.append((j, kind, blk))
    return heading_idx, content


def split_entries(content):
    """把章节内容拆分为 (引导语, 条目列表[(块索引,段落)])。"""
    guide = None
    entries = []
    has_no_ref = False
    for (idx, kind, blk) in content:
        if kind != "p":
            # 表格形式 -> 交给 A4R002 处理
            continue
        txt = blk.text.strip()
        if not txt:
            continue
        if is_no_ref(txt):
            has_no_ref = True
            continue
        if is_guide(txt):
            guide = txt
            continue
        entries.append((idx, blk, txt))
    return guide, entries, has_no_ref


# ---------------------------------------------------------------------------
# 4. 正文引用关系分析工具
# ---------------------------------------------------------------------------
NORMATIVE_VERBS = [
    "应遵循", "应按照", "应符合", "按照", "符合", "应满足", "满足",
    "应达到", "达到", "遵守", "应遵守", "应使用", "使用", "依据", "根据", "引用",
]


def has_normative_verb(text):
    return any(v in text for v in NORMATIVE_VERBS)


def is_note_para(para):
    t = para.text.strip()
    return bool(re.match(r"^注\s*\d*\s*[：:．.\(（]", t)) or t == "注"


def is_source_para(para):
    return "来源：" in para.text or "来源:" in para.text


# ---------------------------------------------------------------------------
# 5. 审核主流程
# ---------------------------------------------------------------------------
def audit(document, rules_by_id):
    blocks = iter_blocks(document)
    heading_idx, content = find_reference_chapter(blocks)

    issues = []  # 每条: dict(rule_id, para, snippet, extra, anchor_idx)

    def add(rule_id, para, snippet="", extra=""):
        rule = rules_by_id.get(rule_id)
        if rule is None:
            return
        issues.append({
            "rule_id": rule_id,
            "rule_name": rule.get("rule_name", ""),
            "category": rule.get("category", ""),
            "severity": rule.get("severity", ""),
            "comment_text": rule.get("comment_text", ""),
            "how_to_fix": rule.get("how_to_fix", ""),
            "para": para,
            "snippet": snippet,
            "extra": extra,
        })

    # ---- A4R001: 章节缺失 ----
    if heading_idx is None:
        first = blocks[0][1] if blocks and blocks[0][0] == "p" else None
        add("A4R001", first, "未找到'规范性引用文件'章节")
        return issues  # 没有章节则后续规则无意义

    heading_para = blocks[heading_idx][1]

    # 章节为表格形式 (A4R002)
    is_table_form = any(kind == "tbl" for (_, kind, _) in content)
    # 章节正文含表格占位符
    placeholder_in_text = any(
        kind == "p" and "$$" in blk.text for (_, kind, blk) in content
    )
    if is_table_form or placeholder_in_text:
        add("A4R002", heading_para, "规范性引用文件以表格形式呈现")
        return issues  # 表格形式无法可靠解析条目, 其余条目级检查跳过

    guide, entries, has_no_ref = split_entries(content)

    # ---- A4R003: 引导语 ----
    if entries:
        if guide is None:
            add("A4R003", heading_para, "存在引用文件但缺少标准引导语")
        if has_no_ref:
            # 有条目却又出现"本文件没有..."
            add("A4R003", heading_para, "存在引用文件却出现'本文件没有规范性引用文件'")
    else:
        if not has_no_ref:
            add("A4R003", heading_para, "无引用文件时未注明'本文件没有规范性引用文件'")

    if not entries:
        # 没有条目, 仅做结构性检查即可
        return issues

    # 解析每个条目的编号
    parsed_entries = []  # dict(idx, para, text, code, prefix, tier, number, year, is_intl)
    for (idx, para, text) in entries:
        raw, code = (None, None)
        found = extract_codes(text, strict=False)
        if found:
            raw, code = found[0]
        parsed_entries.append({
            "idx": idx, "para": para, "text": text,
            "raw": raw, "code": code,
        })

    # ---- A4R005 / A4R006 / A4R007 / A4R008 / A4R009 / A4R015 ----
    _org_prefixes = ["ITU-T", "ITU-R", "ISO/IEC", "IEC", "IEEE", "3GPP"]
    for pe in parsed_entries:
        text = pe["text"]
        code = pe["code"]
        para = pe["para"]

        # A4R007: 国际文件缺少组织代号 (H.xxx 无前缀) —— 与编号是否可解析无关
        if re.search(r"H\.\d+", text) and not any(p in text for p in _org_prefixes):
            add("A4R007", para, text[:60], "国际标准 H.xxx 缺少组织代号")

        # A4R015: 占位符
        if "XXXX" in text.upper() or (code is None and re.search(r"[A-Za-z]+\s*XXXX", text)):
            add("A4R015", para, text[:60], "编号含占位符 XXXX")
            continue  # 占位符条目其它格式规则跳过

        # A4R005: 缺少编号
        if code is None:
            add("A4R005", para, text[:60], "条目无可解析的标准编号")
            continue

        # A4R009: 多余序号/标点
        if re.match(r"^\s*(\d+[.、)]|[（(]\d+[）)])", text) or \
           ("《" in text or "》" in text or "“" in text or "”" in text or '"' in text):
            add("A4R009", para, text[:60], "条目含人工序号或书名号/引号")

        # A4R008: 年份号格式(用连字符而非一字线)
        if re.search(r"\d{3,5}-\d{4}", text):
            add("A4R008", para, text[:60], "年份号使用了 '-' 而非 '—'")

        # A4R006: 国际/国内格式
        tier, _ = classify(code["prefix"])
        if tier == 4:  # 国际
            if not re.search(r"[（(][A-Za-z]", text):
                add("A4R006", para, text[:60], "国际引用文件未以'中文（英文）'结尾")
        else:  # 国内
            if "（英文）" in text or re.search(r"[（(]英文[）)]", text):
                add("A4R006", para, text[:60], "国内引用文件不应以（英文）结尾")

    # ---- A4R004: 排序 ----
    if len(parsed_entries) >= 2:
        order_pairs = []
        ok = True
        for pe in parsed_entries:
            code = pe["code"]
            if code is None:
                continue
            tier, _ = classify(code["prefix"])
            try:
                num = float(code["number"])
            except (TypeError, ValueError):
                num = 0.0
            order_pairs.append((tier, num, pe))
        sorted_pairs = sorted(order_pairs, key=lambda x: (x[0], x[1]))
        if [id(p[2]) for p in order_pairs] != [id(p[2]) for p in sorted_pairs]:
            # 找到第一个失序条目
            first_bad = None
            for i in range(1, len(order_pairs)):
                if (order_pairs[i][0], order_pairs[i][1]) < \
                   (order_pairs[i - 1][0], order_pairs[i - 1][1]):
                    first_bad = order_pairs[i][2]
                    break
            add("A4R004", first_bad["para"] if first_bad else parsed_entries[0]["para"],
                "引用文件未按规范顺序排列")

    # ---- 正文引用关系: A4R010/011/012/013/014 ----
    # 收集条目 base 编码集合
    ref_bases = {}  # base -> pe
    for pe in parsed_entries:
        code = pe["code"]
        if code is None:
            continue
        base = norm_base(code["raw"])
        ref_bases.setdefault(base, pe)

    # 初始化条目使用情况
    usage = {base: {"found": False, "in_note": False, "in_source": False,
                    "normative": False, "note_para": None, "source_para": None,
                    "year_mismatch": []}
             for base in ref_bases}

    # 排除引用章节本身(body 从全文中剔除该章节块)
    exclude_idx = set([heading_idx] + [idx for (idx, _, _) in content])

    a4r014_seen = set()
    for i, (kind, blk) in enumerate(blocks):
        if kind != "p":
            continue
        if i in exclude_idx:
            continue
        para = blk
        text = para.text
        if not text:
            continue
        is_note = is_note_para(para)
        is_source = is_source_para(para)
        normative = has_normative_verb(text)
        codes = extract_codes(text, strict=True)
        for (raw, code) in codes:
            if code is None:
                continue
            base = norm_base(raw)
            if base in usage:
                u = usage[base]
                u["found"] = True
                if is_note:
                    u["in_note"] = True
                    u["note_para"] = para
                if is_source:
                    u["in_source"] = True
                    u["source_para"] = para
                if normative:
                    u["normative"] = True
                # A4R013: 年代号不一致
                entry_pe = ref_bases[base]
                ec = entry_pe["code"]
                if code.get("year") and ec.get("year") and code["year"] != ec["year"]:
                    u["year_mismatch"].append((para, raw, ec.get("raw")))
                elif code.get("year") and not ec.get("year"):
                    u["year_mismatch"].append((para, raw, ec.get("raw")))
            else:
                # 正文中规范性引用但未在第二章列出 -> A4R014
                if normative and base not in a4r014_seen:
                    a4r014_seen.add(base)
                    add("A4R014", para, text[:60], "正文引用 %s 但未在第二章列出" % raw)

    # 条目级引用问题
    for base, u in usage.items():
        pe = ref_bases[base]
        if not u["found"]:
            add("A4R010", pe["para"], pe["text"][:60], "该标准在正文中未被引用")
        elif u["in_note"] and not u["normative"]:
            add("A4R011", u["note_para"] or pe["para"], pe["text"][:60],
                "该标准仅在'注'中提及")
        elif u["in_source"] and not u["normative"]:
            add("A4R012", u["source_para"] or pe["para"], pe["text"][:60],
                "该标准仅在'来源'中提及")
        elif u["found"] and not u["normative"]:
            add("A4R010", pe["para"], pe["text"][:60], "该标准在正文出现但未使用规范性措辞")

    # A4R013: 年代号不一致
    for base, u in usage.items():
        for (para, raw, entry_raw) in u["year_mismatch"]:
            add("A4R013", para, "正文 %s 与条目 %s 年代号不一致" % (raw, entry_raw),
                "正文中引用的标准编号与规范性引用文件条目不一致")

    # 去重(同一段落同一规则只保留一条)
    seen = set()
    deduped = []
    for it in issues:
        key = (it["rule_id"], id(it["para"]), it["snippet"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(it)
    return deduped


# ---------------------------------------------------------------------------
# 6. 插入真·Word 批注
# ---------------------------------------------------------------------------
W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
COMMENTS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments"


def insert_comment_range(para, cid):
    """在段落首尾插入 commentRange 标记(锚定批注到该段落)。"""
    p = para._p
    crs = OxmlElement("w:commentRangeStart")
    crs.set(qn("w:id"), str(cid))
    pPr = p.find(qn("w:pPr"))
    if pPr is not None:
        pPr.addnext(crs)
    else:
        p.insert(0, crs)
    cre = OxmlElement("w:commentRangeEnd")
    cre.set(qn("w:id"), str(cid))
    ref_r = OxmlElement("w:r")
    rPr = OxmlElement("w:rPr")
    rStyle = OxmlElement("w:rStyle")
    rStyle.set(qn("w:val"), "CommentReference")
    rPr.append(rStyle)
    ref_r.append(rPr)
    ref_e = OxmlElement("w:commentReference")
    ref_e.set(qn("w:id"), str(cid))
    ref_r.append(ref_e)
    p.append(cre)
    p.append(ref_r)


def apply_comments(document, issues, author):
    """在文档中插入批注锚点, 返回批注数据列表(供 zip 注入用)。"""
    comments = []
    if not issues:
        return comments
    next_id = 1
    for it in issues:
        para = it["para"]
        if para is None:
            continue
        cid = next_id
        next_id += 1
        body = "[%s] %s\n%s" % (it["rule_id"], it["rule_name"], it["comment_text"])
        if it.get("extra"):
            body += "\n（%s）" % it["extra"]
        comments.append({"id": cid, "text": body})
        insert_comment_range(para, cid)
    return comments


def _now_iso():
    return datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")


def _inject_comment_reference_style(data):
    s = data.decode("utf-8")
    style = (
        '<w:style w:type="character" w:styleId="CommentReference">'
        '<w:name w:val="annotation reference"/></w:style>'
    )
    s = s.replace("</w:styles>", style + "</w:styles>")
    return s.encode("utf-8")


def inject_comments_part(path, comments, author, date):
    """把 comments.xml 与关系注入已保存的 docx(zip), 并补 CommentReference 样式。
    直接读入内存后覆盖写回, 避免临时文件/移动在沙箱中被拦截。"""
    import zipfile
    import re as _re
    from xml.sax.saxutils import escape

    parts = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
             '<w:comments xmlns:w="%s">' % W_NS]
    for c in comments:
        parts.append(
            '<w:comment w:id="%d" w:author="%s" w:date="%s" w:initials="%s">'
            '<w:p><w:r><w:t xml:space="preserve">%s</w:t></w:r></w:p></w:comment>'
            % (c["id"], escape(author), date, escape(author[:1]), escape(c["text"]))
        )
    parts.append("</w:comments>")
    comments_xml = "".join(parts)

    with zipfile.ZipFile(path, "r") as zin:
        entries = [(i, zin.read(i)) for i in zin.namelist()]
        rels = zin.read("word/_rels/document.xml.rels").decode("utf-8")

    ids = [int(m) for m in _re.findall(r'Id="rId(\d+)"', rels)]
    nid = (max(ids) + 1) if ids else 100
    rel_elem = (
        '<Relationship Id="rId%d" Type="%s" Target="comments.xml"/>'
        % (nid, COMMENTS_REL)
    )
    rels = rels.replace("</Relationships>", rel_elem + "</Relationships>")

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries:
            if name == "word/_rels/document.xml.rels":
                continue
            if name == "word/styles.xml" and b"CommentReference" not in data:
                data = _inject_comment_reference_style(data)
            zout.writestr(name, data)
        zout.writestr("word/_rels/document.xml.rels", rels)
        zout.writestr("word/comments.xml", comments_xml)


# ---------------------------------------------------------------------------
# 7. 生成 xlsx 问题清单
# ---------------------------------------------------------------------------
SEVERITY_FILL = {
    "高": PatternFill("solid", fgColor="F4B6B6"),
    "中": PatternFill("solid", fgColor="FCE3B6"),
    "低": PatternFill("solid", fgColor="FFF4C2"),
}


def build_xlsx(issues, out_path, author):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审核问题清单"
    headers = ["序号", "规则编号", "规则名称", "类别", "严重级别",
               "问题描述(批注)", "修改建议", "位置(章节/段落)", "原文片段"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="4472C4")
    head_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for i, it in enumerate(issues, start=1):
        loc = "规范性引用文件"
        ws.append([
            i, it["rule_id"], it["rule_name"], it["category"], it["severity"],
            it["comment_text"], it["how_to_fix"], loc, it.get("snippet", ""),
        ])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        sev_cell = ws.cell(row=r, column=5)
        if it["severity"] in SEVERITY_FILL:
            sev_cell.fill = SEVERITY_FILL[it["severity"]]
            sev_cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [6, 10, 26, 18, 10, 46, 50, 18, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), ws.max_row)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# 8. 主入口
# ---------------------------------------------------------------------------
DEFAULT_RULES_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "rules", "a4guifan_rules.json",
)


def run_normative_audit(input_path, out_docx=None, out_xlsx=None,
                        rules_path=None, author="规范性引用文件审核",
                        write_comments=True, write_xlsx=True):
    """规则驱动的「规范性引用文件」章节审核（A4R001~A4R015）。

    返回 dict：issues / out_docx / out_xlsx / comment_count。
    与 pipeline.py 的启发式引擎互为补充：本模块完全由 rules/a4guifan_rules.json 驱动，
    规则文案（批注内容、修改建议、严重级别）改 JSON 即刻生效，无需改代码。
    """
    rules_path = rules_path or DEFAULT_RULES_PATH
    if not os.path.exists(rules_path):
        raise FileNotFoundError("找不到规则文件 %s" % rules_path)
    _, rules_by_id = load_rules(rules_path)

    if not os.path.exists(input_path):
        raise FileNotFoundError("找不到输入文件 %s" % input_path)

    document = Document(input_path)
    issues = audit(document, rules_by_id)

    base, _ext = os.path.splitext(input_path)
    out_docx = out_docx or (base + "_引用审核批注.docx")
    out_xlsx = out_xlsx or (base + "_引用审核问题.xlsx")

    n_comments = 0
    if write_comments:
        # 写入批注锚点, 保存后再把 comments.xml 注入 zip
        comments = apply_comments(document, issues, author)
        document.save(out_docx)
        if comments:
            inject_comments_part(out_docx, comments, author, _now_iso())
        n_comments = len(comments)
    else:
        out_docx = None

    if write_xlsx:
        build_xlsx(issues, out_xlsx, author)
    else:
        out_xlsx = None

    return {
        "issues": issues,
        "out_docx": out_docx,
        "out_xlsx": out_xlsx,
        "comment_count": n_comments,
    }


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="规范性引用文件自动审核: docx 输入, 输出带批注 docx + 问题 xlsx"
    )
    parser.add_argument("input", help="待审核的 .docx 文件")
    parser.add_argument("-o", "--output-docx", default=None, help="输出带批注 docx (默认 <输入>_引用审核批注.docx)")
    parser.add_argument("-x", "--output-xlsx", default=None, help="输出问题清单 xlsx (默认 <输入>_引用审核问题.xlsx)")
    parser.add_argument("--rules", default=None, help="规则 json 路径 (默认 <仓库>/rules/a4guifan_rules.json)")
    parser.add_argument("--author", default="规范性引用文件审核", help="批注作者名")
    args = parser.parse_args(argv)

    try:
        result = run_normative_audit(
            args.input, args.output_docx, args.output_xlsx,
            args.rules, args.author,
        )
    except FileNotFoundError as exc:
        print("错误: %s" % exc, file=sys.stderr)
        return 2

    print("审核完成: %s" % os.path.basename(args.input))
    print("  发现问题数 : %d" % len(result["issues"]))
    print("  插入批注数 : %d" % result["comment_count"])
    print("  带批注文档 : %s" % result["out_docx"])
    print("  问题清单   : %s" % result["out_xlsx"])
    return 0


if __name__ == "__main__":
    sys.exit(main())
