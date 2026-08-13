#!/usr/bin/env python3
"""
生成两个 xlsx 文件（基于新规则 JSON 审核规则汇总.json）:
1. 审核规则汇总.xlsx — 新规则目录汇总（按模块、审核要点、审核问题、如何修改、批注文本、启用状态）
2. 审核结果_新规则.xlsx — 代码检测到的错误和批注（含规则映射）
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
MODUE_FONT = Font(name='微软雅黑', size=10, bold=True, color='1F4E79')
CELL_FONT = Font(name='微软雅黑', size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top', horizontal='left')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C7E7'),
    right=Side(style='thin', color='B4C7E7'),
    top=Side(style='thin', color='B4C7E7'),
    bottom=Side(style='thin', color='B4C7E7'),
)


def _style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = THIN_BORDER


def _style_cell(cell):
    cell.font = CELL_FONT
    cell.alignment = WRAP_ALIGN
    cell.border = THIN_BORDER


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_rules_summary(rules_path, output_path):
    with open(rules_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    rules = data.get('规则明细', [])
    # 统计信息
    stat = data.get('统计', {})

    wb = Workbook()
    ws = wb.active
    ws.title = '审核规则汇总'

    headers = ['序号', '所属模块', '审核要点', '审核问题', '如何修改', '批注文本(写入文档)', '源码位置', '启用状态']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    # 模块顺序
    module_order = ['主程序(main)', '目次/目录', '前言', '范围', '规范性引用文件',
                    '术语和定义', '缩略语', '正文', '附录']
    rules_sorted = sorted(rules, key=lambda r: (
        module_order.index(r.get('所属模块', '其他')) if r.get('所属模块') in module_order else 99,
        r.get('审核要点', ''),
    ))

    row = 2
    for idx, r in enumerate(rules_sorted, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=r.get('所属模块', ''))
        ws.cell(row=row, column=3, value=r.get('审核要点', ''))
        ws.cell(row=row, column=4, value=r.get('审核问题', ''))
        ws.cell(row=row, column=5, value=r.get('如何修改', ''))
        ws.cell(row=row, column=6, value=r.get('批注文本', ''))
        ws.cell(row=row, column=7, value=r.get('源码位置', ''))
        ws.cell(row=row, column=8, value=r.get('启用状态', ''))
        for col in range(1, 9):
            _style_cell(ws.cell(row=row, column=col))
        ws.cell(row=row, column=2).font = MODUE_FONT
        # 未启用标红
        if '未启用' in r.get('启用状态', ''):
            ws.cell(row=row, column=8).font = Font(name='微软雅黑', size=10, color='C00000')
        row += 1

    _widths(ws, [6, 16, 22, 50, 45, 45, 30, 16])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H{row - 1}'

    # 统计 sheet
    ws2 = wb.create_sheet('统计')
    ws2.cell(row=1, column=1, value='项目')
    ws2.cell(row=1, column=2, value='数值')
    _style_header(ws2, 1, 2)
    stat_rows = [
        ('规则点数(含重复)', stat.get('规则点数(含重复)', '')),
        ('去重后规则数', stat.get('去重后规则数', '')),
        ('按审核要点分类数', stat.get('按审核要点分类数', '')),
        ('已启用规则数', stat.get('已启用规则数', '')),
        ('未启用规则数', stat.get('未启用规则数', '')),
    ]
    for i, (k, v) in enumerate(stat_rows, 2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
        _style_cell(ws2.cell(row=i, column=1))
        _style_cell(ws2.cell(row=i, column=2))

    r = len(stat_rows) + 3
    ws2.cell(row=r, column=1, value='按模块统计 (已启用/总数)')
    ws2.cell(row=r, column=1).font = Font(name='微软雅黑', size=10, bold=True)
    r += 1
    by_module = {}
    for ru in rules:
        m = ru.get('所属模块', '其他')
        by_module.setdefault(m, {'total': 0, 'enabled': 0})
        by_module[m]['total'] += 1
        if 'active' in ru.get('启用状态', '').lower():
            by_module[m]['enabled'] += 1
    for m, c in by_module.items():
        ws2.cell(row=r, column=1, value=m)
        ws2.cell(row=r, column=2, value=f'{c["enabled"]}/{c["total"]}')
        _style_cell(ws2.cell(row=r, column=1))
        _style_cell(ws2.cell(row=r, column=2))
        r += 1
    _widths(ws2, [28, 16])

    wb.save(output_path)
    print(f'审核规则汇总已生成: {output_path} (共 {len(rules)} 条规则明细)')


def generate_review_results(docx_path, rules_path, output_path):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, script_dir)
    from docx_review import DocxAnnotator, NewRuleChecker, generate_results_xlsx

    with open(rules_path, 'r', encoding='utf-8') as f:
        rules_data = json.load(f)

    annotator = DocxAnnotator(docx_path)
    checker = NewRuleChecker(annotator, rules_data)
    checker.check_all()

    annotator.save(output_path.replace('.xlsx', '.docx'))  # 同时生成带批注 docx
    generate_results_xlsx(annotator, checker, output_path)
    print(f'审核结果已生成: {output_path} (共 {len(checker.findings)} 条批注)')
    return len(checker.findings)


def main():
    rules_path = r'C:\Users\hhw82\Desktop\审核2\审核规则汇总.json'
    docx_path = r'C:\Users\hhw82\Desktop\审核\data\2025-0855T-YD_IMS业务保护和容灾技术要求_原始草案报批稿_表格变文字.docx'
    output_dir = r'C:\Users\hhw82\WorkBuddy\文档审核\output'

    os.makedirs(output_dir, exist_ok=True)

    rules_xlsx = os.path.join(output_dir, '审核规则汇总.xlsx')
    generate_rules_summary(rules_path, rules_xlsx)

    results_xlsx = os.path.join(output_dir, '审核结果_新规则.xlsx')
    generate_review_results(docx_path, rules_path, results_xlsx)

    print('\n=== 生成完成 ===')


if __name__ == '__main__':
    main()
