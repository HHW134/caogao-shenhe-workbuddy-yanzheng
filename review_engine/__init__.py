# -*- coding: utf-8 -*-
"""
草案审核 WorkBuddy 验证 —— 统一审核引擎
=======================================

把原先分散在五个工作区的审核方案整合为一个可编排的引擎：

    chapters   术语与引用联动审核  ← 术语章节审核器 + 引用文件审核器 + 审核主控
    normative  规范性引用文件规则审核 ← audit_normative_refs（A4R001~A4R015，规则 JSON 驱动）
    fulldoc    全文档结构审核        ← docx_review（审核规则汇总 156 条，按启用状态执行）

三条链路的结果统一归一化为同一张问题表：
    序号 / 模块 / 审核要点 / 级别 / 问题 / 位置 / 修改建议 / 段落索引 / 原文

对外只暴露一个函数 run_review()，供 CLI（cli.py）与 Web 服务（server.py）共用。
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Iterable, List, Optional

from .terminology import AuditIssue, write_annotated_docx  # noqa: F401
from . import pipeline, normative_refs, fulldoc

__all__ = [
    "MODULES", "run_review", "write_summary_xlsx",
    "AuditIssue", "pipeline", "normative_refs", "fulldoc",
]

# ---------------------------------------------------------------------------
# 模块登记表
# ---------------------------------------------------------------------------

MODULES = {
    "chapters": {
        "name": "术语与引用联动审核",
        "desc": "解析「术语和定义」「规范性引用文件」两章，执行条目级检查并做跨章节联动（术语来源标准不得列入规范性引用文件）。",
        "rules": ["rules/terminology_rules.json", "rules/form_validation_term_ref_rules.json"],
    },
    "normative": {
        "name": "规范性引用文件规则审核",
        "desc": "完全由 rules/a4guifan_rules.json 驱动的 A4R001~A4R015 规则集，改 JSON 即可调整批注文案与级别。",
        "rules": ["rules/a4guifan_rules.json"],
    },
    "fulldoc": {
        "name": "全文档结构审核",
        "desc": "覆盖主程序、目次、范围、规范性引用文件、术语、缩略语、正文等模块，规则来自 rules/fulldoc_rules.json。",
        "rules": ["rules/fulldoc_rules.json"],
    },
}

DEFAULT_MODULES = ("chapters",)

# 所有模块的 docx 批注正文均不再展示「【规则名】[级别]」/「[规则编号] 规则名」/「【审核要点】」
# 这类规则标签前缀，仅保留「问题 / 建议 / 详情」正文，使批注更贴近人工修订意见。
# 该控制分别位于：terminology.py 的 _build_comment_lines()、normative_refs.py 的 apply_comments()、
# fulldoc.py 的 _add()。
SKIP_DOCX_COMMENT_RULES = set()

# 「高/中/低」→ error/warning/info
_SEVERITY_MAP = {"高": "error", "中": "warning", "低": "info"}
_SEVERITY_ORDER = {"error": 0, "warning": 1, "info": 2}


def _severity(raw: str, default: str = "warning") -> str:
    if not raw:
        return default
    raw = str(raw).strip()
    if raw in _SEVERITY_ORDER:
        return raw
    return _SEVERITY_MAP.get(raw, default)


def _para_text(para, limit: int = 80) -> str:
    try:
        text = (para.text or "").strip()
    except AttributeError:
        return ""
    return text[:limit]


# ---------------------------------------------------------------------------
# 各链路执行
# ---------------------------------------------------------------------------

def _run_chapters(docx_path: str, out_dir: Path, stem: str, write_docx: bool):
    result = pipeline.audit_document(docx_path)
    records = []
    for issue in result["all_issues"]:
        d = issue.to_dict()
        records.append({
            "模块": "chapters",
            "模块名称": MODULES["chapters"]["name"],
            "审核要点": d.get("审核要点", ""),
            "级别": _severity(d.get("级别"), "error"),
            "问题": d.get("问题", ""),
            "位置": d.get("位置", ""),
            "修改建议": d.get("修改建议", ""),
            "段落索引": d.get("段落索引", -1),
            "原文": "",
        })

    artifacts = []
    if write_docx:
        out_docx = out_dir / f"{stem}_章节审核批注.docx"
        doc = result["doc"]
        wrote = False
        # 过滤：SKIP_DOCX_COMMENT_RULES 中的规则只进清单，不写 Word 批注
        terms_for_doc = [i for i in result["terms_issues"]
                         if i.rule not in SKIP_DOCX_COMMENT_RULES]
        refs_for_doc = [i for i in result["refs_issues"]
                        if i.rule not in SKIP_DOCX_COMMENT_RULES]
        if (result["para_objects_terms"] or result["title_para_terms"]) and terms_for_doc:
            write_annotated_docx(doc, terms_for_doc, result["para_objects_terms"],
                                 result["title_para_terms"], str(out_docx), author="术语审核")
            wrote = True
        if (result["para_objects_refs"] or result["title_para_refs"]) and refs_for_doc:
            write_annotated_docx(doc, refs_for_doc, result["para_objects_refs"],
                                 result["title_para_refs"], str(out_docx), author="引用审核")
            wrote = True
        if not wrote:
            doc.save(str(out_docx))
        artifacts.append({"类型": "批注文档", "模块": "chapters", "路径": str(out_docx)})

    return records, artifacts, result["meta"]


def _run_normative(docx_path: str, out_dir: Path, stem: str, write_docx: bool):
    out_docx = out_dir / f"{stem}_引用规则审核批注.docx" if write_docx else None
    result = normative_refs.run_normative_audit(
        docx_path,
        out_docx=str(out_docx) if out_docx else None,
        write_comments=write_docx,
        write_xlsx=False,
    )
    records = []
    for issue in result["issues"]:
        location = issue.get("snippet") or _para_text(issue.get("para"))
        records.append({
            "模块": "normative",
            "模块名称": MODULES["normative"]["name"],
            "审核要点": f"{issue.get('rule_id', '')} {issue.get('rule_name', '')}".strip(),
            "级别": _severity(issue.get("severity")),
            "问题": issue.get("comment_text", ""),
            "位置": location or issue.get("category", ""),
            "修改建议": issue.get("how_to_fix", ""),
            "段落索引": -1,
            "原文": _para_text(issue.get("para")),
        })

    artifacts = []
    if result.get("out_docx"):
        artifacts.append({"类型": "批注文档", "模块": "normative", "路径": result["out_docx"]})
    meta = {"批注数": result.get("comment_count", 0)}
    return records, artifacts, meta


def _run_fulldoc(docx_path: str, out_dir: Path, stem: str, write_docx: bool):
    out_docx = out_dir / f"{stem}_全文档审核批注.docx" if write_docx else None
    result = fulldoc.run_fulldoc_review(
        docx_path,
        out_docx=str(out_docx) if out_docx else None,
        write_comments=write_docx,
        write_xlsx=False,
    )
    records = []
    for f in result["findings"]:
        records.append({
            "模块": "fulldoc",
            "模块名称": MODULES["fulldoc"]["name"],
            "审核要点": f"{f.get('module', '')}-{f.get('audit_point', '')}".strip("-"),
            "级别": "warning",
            "问题": f.get("comment", ""),
            "位置": f"段落 {f.get('para_idx', -1)}",
            "修改建议": f.get("rule_question", "") or "参见规则明细「如何修改」",
            "段落索引": f.get("para_idx", -1),
            "原文": f.get("orig_text", ""),
        })

    artifacts = []
    if result.get("out_docx"):
        artifacts.append({"类型": "批注文档", "模块": "fulldoc", "路径": result["out_docx"]})
    return records, artifacts, result.get("rule_stats", {})


_RUNNERS = {
    "chapters": _run_chapters,
    "normative": _run_normative,
    "fulldoc": _run_fulldoc,
}


# ---------------------------------------------------------------------------
# 统一入口
# ---------------------------------------------------------------------------

def run_review(docx_path,
               modules: Iterable[str] = DEFAULT_MODULES,
               out_dir=None,
               write_docx: bool = True,
               write_xlsx: bool = True) -> dict:
    """对一份 docx 执行选定模块的审核，返回统一结果字典。

    参数
    ----
    docx_path : 待审核的 .docx
    modules   : chapters / normative / fulldoc 的任意组合
    out_dir   : 产物输出目录（默认与源文件同目录）
    write_docx: 是否生成带 Word 批注的文档
    write_xlsx: 是否生成统一问题清单 xlsx
    """
    docx_path = str(docx_path)
    src = Path(docx_path)
    if not src.exists():
        raise FileNotFoundError(f"找不到输入文件：{docx_path}")

    modules = [m for m in modules if m in _RUNNERS] or list(DEFAULT_MODULES)
    out_dir = Path(out_dir) if out_dir else src.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = src.stem

    all_records: List[dict] = []
    artifacts: List[dict] = []
    module_meta = {}
    errors = {}

    for mod in modules:
        try:
            records, arts, meta = _RUNNERS[mod](docx_path, out_dir, stem, write_docx)
        except Exception as exc:  # 单模块失败不影响其余模块
            errors[mod] = f"{type(exc).__name__}: {exc}"
            continue
        all_records.extend(records)
        artifacts.extend(arts)
        module_meta[mod] = meta

    all_records.sort(key=lambda r: (_SEVERITY_ORDER.get(r["级别"], 9), r["模块"]))
    for i, rec in enumerate(all_records, 1):
        rec["序号"] = i

    stats = {"error": 0, "warning": 0, "info": 0}
    by_module = {}
    for rec in all_records:
        stats[rec["级别"]] = stats.get(rec["级别"], 0) + 1
        by_module[rec["模块名称"]] = by_module.get(rec["模块名称"], 0) + 1

    summary = {
        "文件名": src.name,
        "审核时间": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "执行模块": [MODULES[m]["name"] for m in modules if m in MODULES],
        "模块标识": modules,
        "问题总数": len(all_records),
        "级别统计": stats,
        "模块统计": by_module,
        "章节信息": module_meta.get("chapters", {}),
        "模块详情": module_meta,
        "审核结论": "通过" if not all_records else "存在问题",
        "问题列表": all_records,
        "产物": artifacts,
    }
    if errors:
        summary["模块异常"] = errors

    if write_xlsx:
        xlsx_path = out_dir / f"{stem}_审核问题清单.xlsx"
        write_summary_xlsx(summary, str(xlsx_path))
        artifacts.append({"类型": "问题清单", "模块": "all", "路径": str(xlsx_path)})

    return summary


# ---------------------------------------------------------------------------
# 统一 xlsx 汇总
# ---------------------------------------------------------------------------

def write_summary_xlsx(summary: dict, out_path: str) -> str:
    """把统一问题表导出为 xlsx（概览 + 审核问题两个工作表）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    cell_font = Font(name="微软雅黑", size=10)
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
    thin = Side(style="thin", color="B4C7E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "error": PatternFill("solid", start_color="FCE4E4", end_color="FCE4E4"),
        "warning": PatternFill("solid", start_color="FFF6E0", end_color="FFF6E0"),
        "info": PatternFill("solid", start_color="E7F1FB", end_color="E7F1FB"),
    }

    wb = Workbook()

    ws = wb.active
    ws.title = "概览"
    stats = summary.get("级别统计", {})
    rows = [
        ("文件名", summary.get("文件名", "")),
        ("审核时间", summary.get("审核时间", "")),
        ("执行模块", "、".join(summary.get("执行模块", []))),
        ("问题总数", summary.get("问题总数", 0)),
        ("错误 error", stats.get("error", 0)),
        ("警告 warning", stats.get("warning", 0)),
        ("提示 info", stats.get("info", 0)),
        ("审核结论", summary.get("审核结论", "")),
    ]
    for name, count in (summary.get("模块统计") or {}).items():
        rows.append((f"模块问题数：{name}", count))
    chapter_info = summary.get("章节信息") or {}
    for chapter, info in chapter_info.items():
        if isinstance(info, dict) and info:
            detail = "；".join(f"{k}={v}" for k, v in info.items()
                               if not isinstance(v, (list, dict)))
            rows.append((f"章节：{chapter}", detail))
    for r, (k, v) in enumerate(rows, 1):
        c1, c2 = ws.cell(row=r, column=1, value=k), ws.cell(row=r, column=2, value=v)
        c1.font = Font(name="微软雅黑", size=10, bold=True)
        c2.font = cell_font
        for c in (c1, c2):
            c.alignment = wrap
            c.border = border
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 80

    ws2 = wb.create_sheet("审核问题")
    headers = ["序号", "模块", "审核要点", "级别", "问题", "位置", "修改建议", "段落索引", "原文"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, wrap, border
    for r, rec in enumerate(summary.get("问题列表", []), 2):
        values = [
            rec.get("序号", r - 1), rec.get("模块名称", rec.get("模块", "")),
            rec.get("审核要点", ""), rec.get("级别", ""), rec.get("问题", ""),
            rec.get("位置", ""), rec.get("修改建议", ""), rec.get("段落索引", -1),
            rec.get("原文", ""),
        ]
        for col, v in enumerate(values, 1):
            c = ws2.cell(row=r, column=col, value=v)
            c.font, c.alignment, c.border = cell_font, wrap, border
            fill = fills.get(rec.get("级别"))
            if fill:
                c.fill = fill
    for col, width in enumerate([6, 20, 26, 10, 46, 30, 46, 10, 30], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = "A2"
    if summary.get("问题列表"):
        ws2.auto_filter.ref = f"A1:I{len(summary['问题列表']) + 1}"

    wb.save(out_path)
    return out_path
